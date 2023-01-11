import datetime
from nepali.datetime import nepalidate
import openpyxl as xl
from openpyxl.styles import PatternFill

wb = xl.load_workbook('data_log.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']

np_date = nepalidate.today()
date = datetime.date.today()

print(date)
row = 2

for year in range(2073, 1999, -1):
    for month in range(1, 13):
        cell_nepali_year = sheet.cell(row, 1)
        cell_nepali_month = sheet.cell(row, 2)
        cell_english_date = sheet.cell(row, 3)
        cell_nepali_year.value = year
        cell_nepali_month.value = month
        row = row + 1
        np_date = nepalidate(year, month, 1)
        date = np_date.to_date()
        cell_english_date.value = date
    row = row + 2
    a = sheet.cell(row, 1)
    b = sheet.cell(row, 2)
    c = sheet.cell(row, 3)
    a.fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    a.value = 'Nepali Year'
    b.fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    b.value = 'Nepali Month'
    c.fill = PatternFill(start_color="dddddd", end_color="dddddd", fill_type="solid")
    c.value = 'English Date'
    row = row + 1

wb.save('date_log.xlsx')

